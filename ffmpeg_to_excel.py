#!/usr/bin/env python3
"""
Convert FFmpeg output file to Excel spreadsheet
Extracts specific fields from [FRAME]...[/FRAME] blocks
"""

import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import os
import argparse

def parse_ffmpeg_output(file_path, skip_audio=True):
    """Parse FFmpeg output file and extract frame data
    
    Args:
        file_path: Path to the input file
        skip_audio: If True, skip frames with media_type=audio (default: True)
    """
    
    print(f"Reading file: {file_path}")
    
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
    except Exception as e:
        print(f"Error reading file: {e}")
        return None
    
    print(f"File loaded successfully. Size: {len(content)} characters")
    
    # Find all frame blocks
    frame_pattern = r'\[FRAME\](.*?)\[/FRAME\]'
    frame_blocks = re.findall(frame_pattern, content, re.DOTALL)
    
    print(f"Found {len(frame_blocks)} frame blocks")
    
    # Define the fields we want to extract
    target_fields = [
        'media_type',
        'pts', 
        'pts_time',
        'pkt_dts',
        'pkt_dts_time', 
        'best_effort_timestamp',
        'best_effort_timestamp_time',
        'pict_type',
        'duration',
        'duration_time',
        'key_frame',
        'stream_index'
    ]
    
    # Initialize data structure
    frame_data = []
    skipped_audio_count = 0
    
    for i, block in enumerate(frame_blocks):
        if i % 1000 == 0:
            print(f"Processing frame {i+1}/{len(frame_blocks)}")
        
        # Check if this is an audio frame and should be skipped
        if skip_audio:
            media_type_match = re.search(r'^media_type=(.*)$', block, re.MULTILINE)
            if media_type_match and media_type_match.group(1).strip() == 'audio':
                skipped_audio_count += 1
                continue
            
        frame_info = {}
        
        # Extract each field from the block
        for field in target_fields:
            # Create regex pattern for field=value
            pattern = rf'^{re.escape(field)}=(.*)$'
            match = re.search(pattern, block, re.MULTILINE)
            
            if match:
                value = match.group(1).strip()
                # Convert numeric values
                if field in ['pts', 'pkt_dts', 'best_effort_timestamp', 'duration', 'key_frame', 'stream_index']:
                    try:
                        frame_info[field] = int(value) if value != 'N/A' else None
                    except ValueError:
                        frame_info[field] = value
                elif field in ['pts_time', 'pkt_dts_time', 'best_effort_timestamp_time', 'duration_time']:
                    try:
                        frame_info[field] = float(value) if value != 'N/A' else None
                    except ValueError:
                        frame_info[field] = value
                else:
                    frame_info[field] = value
            else:
                frame_info[field] = None
        
        frame_data.append(frame_info)
    
    if skip_audio and skipped_audio_count > 0:
        print(f"Skipped {skipped_audio_count} audio frames")
    
    return frame_data

def create_excel_file(frame_data, output_path):
    """Create Excel file from frame data"""
    
    print(f"Creating Excel file with {len(frame_data)} rows")
    
    # Create DataFrame
    df = pd.DataFrame(frame_data)
    
    # Reorder columns to match the requested order
    column_order = [
        'media_type',
        'pts', 
        'pts_time',
        'pkt_dts',
        'pkt_dts_time', 
        'best_effort_timestamp',
        'best_effort_timestamp_time',
        'pict_type',
        'duration',
        'duration_time',
        'key_frame',
        'stream_index'
    ]
    
    # Only include columns that exist in the data
    existing_columns = [col for col in column_order if col in df.columns]
    df = df[existing_columns]
    
    # Save to Excel
    try:
        df.to_excel(output_path, index=False, engine='openpyxl')
        
        # Add delta columns with formulas
        wb = load_workbook(output_path)
        ws = wb.active
        
        # Add headers for delta columns (after the existing columns)
        last_col = ws.max_column
        ws.cell(row=1, column=last_col + 1, value='pts_delta')
        ws.cell(row=1, column=last_col + 2, value='pts_time_delta')
        ws.cell(row=1, column=last_col + 3, value='pkt_dts_delta')
        ws.cell(row=1, column=last_col + 4, value='pkt_dts_time_delta')
        
        # Add formulas starting from row 3 (second data row, skipping first data row)
        # Row 1 is header, Row 2 is first data row (skip), Row 3+ get formulas
        for row_num in range(3, ws.max_row + 1):
            # pts_delta: =B(N)-B(N-1)
            ws.cell(row=row_num, column=last_col + 1, value=f'=B{row_num}-B{row_num-1}')
            # pts_time_delta: =C(N)-C(N-1)
            ws.cell(row=row_num, column=last_col + 2, value=f'=C{row_num}-C{row_num-1}')
            # pkt_dts_delta: =D(N)-D(N-1)
            ws.cell(row=row_num, column=last_col + 3, value=f'=D{row_num}-D{row_num-1}')
            # pkt_dts_time_delta: =E(N)-E(N-1)
            ws.cell(row=row_num, column=last_col + 4, value=f'=E{row_num}-E{row_num-1}')
        
        # Save the workbook with formulas
        wb.save(output_path)
        print(f"Excel file saved successfully: {output_path}")
        print(f"Added delta columns with formulas for rows 3-{ws.max_row}")
        
        # Print summary statistics
        print(f"\nSummary:")
        print(f"Total frames: {len(df)}")
        if 'media_type' in df.columns:
            print(f"Media type breakdown:")
            print(df['media_type'].value_counts())
        if 'stream_index' in df.columns:
            print(f"Stream index breakdown:")
            print(df['stream_index'].value_counts())
            
    except Exception as e:
        print(f"Error saving Excel file: {e}")
        return False
    
    return True

def main():
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description='Convert FFmpeg output file to Excel spreadsheet',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python ffmpeg_to_excel.py output.txt
  python ffmpeg_to_excel.py /path/to/ffmpeg_output.txt -o frames.xlsx
        '''
    )
    
    parser.add_argument('input_file', 
                       help='Path to the FFmpeg output file to convert')
    parser.add_argument('-o', '--output', 
                       help='Output Excel file path (default: based on input filename)')
    parser.add_argument('--include-audio', 
                       action='store_true',
                       help='Include audio frames (by default, audio frames are skipped)')
    
    args = parser.parse_args()
    
    # Input file validation
    input_file = args.input_file
    if not os.path.exists(input_file):
        print(f"Error: Input file '{input_file}' does not exist")
        return 1
    
    # Generate output filename if not provided
    if args.output:
        output_file = args.output
    else:
        # Create output filename based on input filename
        input_name = os.path.splitext(os.path.basename(input_file))[0]
        output_file = f"{input_name}_frames.xlsx"
    
    # Ensure output directory exists
    output_dir = os.path.dirname(output_file)
    if output_dir and not os.path.exists(output_dir):
        try:
            os.makedirs(output_dir)
        except Exception as e:
            print(f"Error creating output directory: {e}")
            return 1
    
    print("=== FFmpeg Output to Excel Converter ===\n")
    print(f"Input file: {input_file}")
    print(f"Output file: {output_file}")
    print(f"Skip audio frames: {not args.include_audio}\n")
    
    # Parse the FFmpeg output
    frame_data = parse_ffmpeg_output(input_file, skip_audio=not args.include_audio)
    
    if frame_data is None:
        print("Failed to parse input file")
        return 1
    
    if len(frame_data) == 0:
        print("No frame data found")
        return 1
    
    # Create Excel file
    success = create_excel_file(frame_data, output_file)
    
    if success:
        print(f"\nConversion completed successfully!")
        print(f"Input file: {input_file}")
        print(f"Output file: {output_file}")
        return 0
    else:
        print("Conversion failed")
        return 1

if __name__ == "__main__":
    sys.exit(main())
